# Excel3.py
import openpyxl

book = openpyxl.load_workbook("/Users/devante/Library/CloudStorage/OneDrive-ECPIUniversity/Desktop/dDominicci_cis_123/SecondBook.xlsx")
sheet = book.active
rows = sheet.rows
book.close()
values = []
for row in rows:
    for cell in row:
        value = (cell.value)
        if cell.column == 2:
            value = value * 10
        values.append(value)
book = openpyxl.Workbook()
sheet = book.active
index = 0
rw = 1
col = 1
for val in values:
    sheet.cell(row=rw, column=col).value = val
    index += 1
    col += 1
    if index == 3:
        rw += 1
        col = 1
        index = 0
book.save("/Users/devante/Library/CloudStorage/OneDrive-ECPIUniversity/Desktop/dDominicci_cis_123/ThirdBook.xlsx")
